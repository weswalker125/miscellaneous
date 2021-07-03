import os
import sys
import pandas
from shutil import copyfile
import datetime
import requests
from tqdm import tqdm
from openpyxl import load_workbook
import subprocess

# Primary working location 
rootDir = "/tmp/zack"
templateExcelFile = rootDir + "/" + "template.xlsm"
today = datetime.date.today()
workingDir = rootDir + "/" + today.strftime("%Y-%m-%d")
urlFormat = "https://www.zacks.com/{service_name}/download_trades.php"
headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; rv:89.0) Gecko/20100101 Firefox/89.0" }
downloadedFiles = []

# Names of services available to Zack's subscribers
services = [
    "etfinvestor",
    "homerun",
    "incomeinvestor",
    "valueinvestor",
    "top10",
    "blackboxtrader",
    "counterstrike",
    "headlinetrader",
    "insidertrader",
    "largecaptrader",
    "optionstrader",
    "shortlist",
    "surprisetrader",
    "tazr"
]

# Create working directory for today
os.makedirs(workingDir, exist_ok=True)

# Copy template
if not os.path.isfile(templateExcelFile):
    print("No template found at default location [{}]".format(templateExcelFile))
    templateExcelFile = input("Template file location: ")
try:
    mainExcelFile = workingDir + "/" + "main.xlsx"
    copyfile(templateExcelFile, mainExcelFile)
except Exception as err:
    print("Unable to open template file: " + templateExcelFile)
    print(err)
    sys.exit(1)


# Download the trades
for service in services:
    url = urlFormat.format(service_name = service)
    try:
        filename = workingDir + "/" + service + ".xlsx"
        # Download the file if we don't have it
        if not os.path.isfile(filename):
            print("Downloading from " + url)
            r = requests.get(url, headers=headers, allow_redirects=True, stream=True)
            with open(filename, "wb") as f:
                for data in tqdm(r.iter_content()):
                    f.write(data)
        downloadedFiles.append({ "name": service, "file": filename })
    except Exception as err:
        print("Error occurred downloading for " + service)
        print(err)

# Open the main workbook (copied from the template)
mainWorkbook = load_workbook(filename=mainExcelFile)
# Copy the first sheet of each downloaded file into the main workbook
for f in downloadedFiles:
    try:
        # load workbook
        wb = load_workbook(f["file"])
        sheet = wb.worksheets[0]

        # Add sheet to main file
        newSheet = mainWorkbook.create_sheet(f["name"])

        # Copy data from each row in sheet
        for row in sheet:
            for cell in row:
                newSheet[cell.coordinate].value = cell.value
    except Exception as err:
        print("Failed to read file: " + f["file"])
        print(err)

try:
    mainWorkbook.save(mainExcelFile)
    print("Today's workbook is at " + mainExcelFile)

    print("Attempting to open workbook...")
    subprocess.run(["open", mainExcelFile])
    sys.exit(0)
except Exception as err:
    print("Unable to save workbook")
    print(err)
    sys.exit(2)